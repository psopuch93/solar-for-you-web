from rest_framework import viewsets, permissions, status, parsers
from rest_framework.decorators import action, api_view, permission_classes
from django.http import JsonResponse
from django.views.decorators.csrf import ensure_csrf_cookie
from rest_framework.response import Response
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login
from django.contrib.auth.models import User
from .utils.email_utils import send_requisition_notification
from django.db.models import Q
from django.utils.decorators import method_decorator
import datetime
import json
from .models import UserProfile, Project, Client, ProjectTag, Employee, Empl_tag, Requisition, Item, RequisitionItem, Quarter, QuarterImage, UserSettings, BrigadeMember,HRRequisitionPosition, HRRequisition, TransportRequest, TransportItem
from .serializers import (
    UserSerializer, UserProfileSerializer, ProjectSerializer,
    ClientSerializer, ProjectTagSerializer, EmployeeSerializer, EmplTagSerializer,
    ItemSerializer, RequisitionSerializer, RequisitionItemSerializer, QuarterSerializer, QuarterImageSerializer,
    UserSettingsSerializer, BrigadeMemberSerializer, ProgressReportSerializer,
    ProgressReport, ProgressReportEntrySerializer, ProgressReportEntry, ProgressReportImageSerializer,
    ProgressReportImage, HRRequisitionPositionSerializer, HRRequisitionSerializer, TransportRequestSerializer, TransportItemSerializer,
    ProgressReportActivitySerializer, ProjectActivityConfig, ProgressReportActivity, ProjectActivityConfigSerializer
)

class IsAdminOrOwner(permissions.BasePermission):
    """
    Custom permission to only allow owners of an object or admins to edit it.
    """
    def has_object_permission(self, request, view, obj):
        # Read permissions are allowed to any request
        if request.method in permissions.SAFE_METHODS:
            return True

        # Admin zawsze ma dostęp
        if request.user.is_staff:
            return True

        # Owner
        if hasattr(obj, 'user'):
            return obj.user == request.user
        elif hasattr(obj, 'client'):
            return obj.client == request.user
        elif hasattr(obj, 'author'):
            return obj.author == request.user
        else:
            return False

class HasModulePrivilege(permissions.BasePermission):
    """
    Custom permission do sprawdzania uprawnień do konkretnego modułu.
    Wymagane uprawnienie należy podać w view.required_privilege.
    """
    message = "Brak uprawnień do tego modułu."

    def has_permission(self, request, view):
        # Admin zawsze ma dostęp
        if request.user.is_staff:
            return True

        # Pobieramy wymagane uprawnienie z widoku
        required_privilege = getattr(view, 'required_privilege', None)
        if not required_privilege:
            return True  # Jeśli nie określono uprawnienia, domyślnie pozwalamy

        # Sprawdzamy uprawnienia użytkownika
        try:
            profile = request.user.profile
            return profile.has_privilege(required_privilege)
        except UserProfile.DoesNotExist:
            return False

    def has_object_permission(self, request, view, obj):
        # Dla bezpieczeństwa sprawdzamy również na poziomie obiektu
        return self.has_permission(request, view)

class UserViewSet(viewsets.ModelViewSet):
    """API endpoint dla użytkowników"""
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [permissions.IsAdminUser]
    required_privilege = 'admin_users'  # Uprawnienie do zarządzania użytkownikami

    @action(detail=False, methods=['get'], permission_classes=[permissions.IsAuthenticated])
    def me(self, request):
        """Endpoint zwracający dane zalogowanego użytkownika"""
        serializer = self.get_serializer(request.user)
        return Response(serializer.data)

class UserProfileViewSet(viewsets.ModelViewSet):
    """API endpoint dla profili użytkowników"""
    queryset = UserProfile.objects.all()
    serializer_class = UserProfileSerializer
    permission_classes = [IsAdminOrOwner, HasModulePrivilege]
    required_privilege = 'manage_users'  # Uprawnienie do zarządzania profilami

    @action(detail=False, methods=['get'], permission_classes=[permissions.IsAuthenticated])
    def my_profile(self, request):
        """Endpoint zwracający profil zalogowanego użytkownika"""
        profile, created = UserProfile.objects.get_or_create(user=request.user)
        serializer = self.get_serializer(profile)
        return Response(serializer.data)

class ProjectViewSet(viewsets.ModelViewSet):
    """API endpoint dla projektów"""
    queryset = Project.objects.all()
    serializer_class = ProjectSerializer
    permission_classes = [IsAdminOrOwner, HasModulePrivilege]
    required_privilege = 'manage_projects'  # Uprawnienie do zarządzania projektami

    def get_queryset(self):
        """Filtrowanie projektów w zależności od uprawnień użytkownika"""
        user = self.request.user

        # Admin widzi wszystkie projekty
        if user.is_staff:
            return Project.objects.all()

        # Sprawdzamy uprawnienia użytkownika
        try:
            profile = user.profile
            # Jeśli użytkownik ma uprawnienie do zarządzania wszystkimi projektami
            if profile.has_privilege('view_all_projects'):
                return Project.objects.all()
            # W przeciwnym razie widzi tylko swoje projekty
            return Project.objects.filter(client=user)
        except UserProfile.DoesNotExist:
            # Domyślnie zwracamy projekty, których użytkownik jest klientem
            return Project.objects.filter(client=user)

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user, updated_by=self.request.user)

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)

class ClientViewSet(viewsets.ModelViewSet):
    """API endpoint dla klientów"""
    queryset = Client.objects.all()
    serializer_class = ClientSerializer
    permission_classes = [IsAdminOrOwner, HasModulePrivilege]
    required_privilege = 'manage_clients'  # Nowe uprawnienie

    def perform_create(self, serializer):
        """Automatycznie ustaw użytkownika tworzącego klienta"""
        serializer.save(created_by=self.request.user)

    def perform_update(self, serializer):
        """Automatycznie ustaw użytkownika aktualizującego klienta"""
        serializer.save(updated_by=self.request.user)

    def get_queryset(self):
        """Filtrowanie klientów"""
        user = self.request.user

        # Admin widzi wszystkich klientów
        if user.is_staff:
            return Client.objects.all()

        # Domyślnie zwracamy pustą listę dla zwykłego użytkownika
        return Client.objects.none()

class ProjectTagViewSet(viewsets.ModelViewSet):
    """API endpoint dla tagów projektów"""
    queryset = ProjectTag.objects.all()
    serializer_class = ProjectTagSerializer
    permission_classes = [permissions.IsAuthenticated, HasModulePrivilege]
    required_privilege = 'manage_project_tags'  # Uprawnienie do zarządzania tagami

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user, updated_by=self.request.user)

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)

# Istniejące widoki logowania, dashboard itp. pozostają bez zmian
from django.views.decorators.http import require_http_methods

@require_http_methods(["GET", "POST"])
def login_api(request):
    """API do logowania"""
    if request.method == 'POST':
        # Obsługa żądań JSON
        try:
            data = json.loads(request.body)
            username = data.get('username')
            password = data.get('password')
        except json.JSONDecodeError:
            username = request.POST.get('username')
            password = request.POST.get('password')

        if not username or not password:
            return JsonResponse({
                'success': False,
                'message': 'Proszę podać nazwę użytkownika i hasło'
            }, status=400)

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            return JsonResponse({
                'success': True,
                'message': 'Zalogowano pomyślnie',
                'redirect': '/dashboard/'
            })
        else:
            return JsonResponse({
                'success': False,
                'message': 'Nieprawidłowa nazwa użytkownika lub hasło'
            }, status=401)

    # Dla metody GET - renderowanie widoku logowania
    return render(request, 'index.html')

def login_view(request):
    """Widok renderujący stronę logowania"""
    # Jeśli użytkownik jest już zalogowany, przekieruj do dashboardu
    if request.user.is_authenticated:
        return redirect('dashboard')

    return render(request, 'index.html')

@login_required
def dashboard_view(request):
    """Widok renderujący dashboard React (wymaga logowania)"""
    return render(request, 'dashboard.html')

@require_http_methods(["POST"])
def logout_api(request):
    """API do wylogowania"""
    from django.contrib.auth import logout
    logout(request)
    return JsonResponse({
        'success': True,
        'message': 'Wylogowano pomyślnie'
    })

# Nowy endpoint do sprawdzania unikalności nazwy projektu
@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def check_project_name(request):
    # Checks if a project name is unique
    name = request.GET.get('name')
    project_id = request.GET.get('id')  # Optional, for editing

    if not name:
        return Response({'valid': False, 'message': 'No project name provided'}, status=status.HTTP_400_BAD_REQUEST)

    # Check if name has minimum length
    if len(name) < 3:
        return Response({'valid': False, 'message': 'Project name must be at least 3 characters'})

    # For editing case - exclude current project
    query = Q(name=name)
    if project_id:
        query &= ~Q(id=project_id)

    exists = Project.objects.filter(query).exists()

    if exists:
        return Response({'valid': False, 'message': 'A project with this name already exists'})
    else:
        return Response({'valid': True, 'message': 'Name is available'})

# Add these ViewSets to the existing views.py file

class EmplTagViewSet(viewsets.ModelViewSet):
    """API endpoint dla tagów pracowników"""
    queryset = Empl_tag.objects.all()
    serializer_class = EmplTagSerializer
    permission_classes = [permissions.IsAuthenticated, HasModulePrivilege]
    required_privilege = 'manage_employee_tags'  # Uprawnienie do zarządzania tagami

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user, updated_by=self.request.user)

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)

class EmployeeViewSet(viewsets.ModelViewSet):
    """API endpoint dla pracowników"""
    queryset = Employee.objects.all()
    serializer_class = EmployeeSerializer
    permission_classes = [permissions.IsAuthenticated, HasModulePrivilege]
    required_privilege = 'manage_employees'  # Uprawnienie do zarządzania pracownikami

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user, updated_by=self.request.user)

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)

# New endpoint to check PESEL uniqueness
@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def check_pesel(request):
    """Sprawdza czy PESEL jest unikalny"""
    pesel = request.GET.get('pesel')
    employee_id = request.GET.get('id')  # Optional, for editing

    # Jeśli PESEL jest pusty, zwróć od razu true (pusty PESEL jest dozwolony)
    if not pesel or pesel.strip() == '':
        return Response({'valid': True, 'message': 'PESEL nie jest wymagany'})

    # Sprawdzamy czy PESEL ma odpowiednią długość i zawiera tylko cyfry
    if len(pesel) != 11 or not pesel.isdigit():
        return Response({'valid': False, 'message': 'PESEL musi składać się z 11 cyfr'})

    # Dla przypadku edycji - wykluczamy bieżącego pracownika
    query = Q(pesel=pesel)
    if employee_id:
        query &= ~Q(id=employee_id)

    exists = Employee.objects.filter(query).exists()

    if exists:
        return Response({'valid': False, 'message': 'Pracownik o tym numerze PESEL już istnieje'})
    else:
        return Response({'valid': True, 'message': 'PESEL dostępny'})

class ItemViewSet(viewsets.ModelViewSet):
    """API endpoint dla przedmiotów"""
    queryset = Item.objects.all()
    serializer_class = ItemSerializer
    permission_classes = [permissions.IsAuthenticated, HasModulePrivilege]
    required_privilege = 'manage_items'

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user, updated_by=self.request.user)

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)


@method_decorator(ensure_csrf_cookie, name='dispatch')
class RequisitionViewSet(viewsets.ModelViewSet):
    """API endpoint dla zapotrzebowań"""
    queryset = Requisition.objects.all()
    serializer_class = RequisitionSerializer
    permission_classes = [permissions.IsAuthenticated, HasModulePrivilege]
    required_privilege = 'manage_requisitions'

    def get_queryset(self):
        """Filtruj zapotrzebowania z obsługą wyszukiwania po przedmiotach"""
        user = self.request.user
        # Użytkownicy z uprawnieniem 'view_all_requisitions' lub admin mogą widzieć wszystkie zapotrzebowania
        if user.is_staff or hasattr(user, 'profile') and user.profile.has_privilege('view_all_requisitions'):
            queryset = Requisition.objects.all().order_by('-created_at')
        else:
            # Pozostali użytkownicy widzą tylko swoje zapotrzebowania
            queryset = Requisition.objects.filter(created_by=user).order_by('-created_at')

        # Filtruj po typie zapotrzebowania
        requisition_type = self.request.query_params.get('requisition_type', None)
        if requisition_type:
            queryset = queryset.filter(requisition_type=requisition_type)

        # Filtruj po frazie wyszukiwania
        search_term = self.request.query_params.get('search', None)
        if search_term:
            # Wyszukiwanie w podstawowych polach zapotrzebowania
            basic_search = Q(number__icontains=search_term) | \
                           Q(comment__icontains=search_term)

            # Dodaj wyszukiwanie w polach powiązanego projektu
            basic_search |= Q(project__name__icontains=search_term)

            # Wyszukiwanie w przedmiotach zapotrzebowania
            items_search = RequisitionItem.objects.filter(
                Q(item__name__icontains=search_term) |
                Q(item__index__icontains=search_term)
            ).values_list('requisition_id', flat=True).distinct()

            # Łączymy oba warunki - podstawowe pola lub powiązane przedmioty
            queryset = queryset.filter(basic_search | Q(id__in=items_search))

        return queryset.distinct()

    def get_serializer_context(self):
        """Dodaj request do kontekstu serializera, aby mieć dostęp do aktualnego użytkownika"""
        context = super().get_serializer_context()
        context['request'] = self.request
        return context

    def perform_create(self, serializer):
        """
        Modyfikacja procesu tworzenia zapotrzebowania,
        obsługująca wymuszenie nowego numeru
        """
        requisition = serializer.save(
            created_by=self.request.user,
            updated_by=self.request.user
        )

        # Wysyłanie powiadomienia e-mail
        try:
            send_requisition_notification(requisition)
        except Exception:
            # Usunięto logowanie błędu
            pass

    def perform_update(self, serializer):
        """
        Obsługa aktualizacji zapotrzebowania z dodatkową walidacją
        """
        # Pobierz aktualny stan przed aktualizacją
        instance = self.get_object()
        old_status = instance.status

        # Zaktualizuj zapotrzebowanie
        updated_instance = serializer.save(
            updated_by=self.request.user
        )

        # Opcjonalnie: wyślij powiadomienie o zmianie statusu
        if old_status != updated_instance.status:
            try:
                # Możesz zdefiniować osobną funkcję wysyłania powiadomień o zmianie statusu
                # send_status_change_notification(updated_instance)
                pass
            except Exception:
                # Usunięto logowanie błędu
                pass

    def partial_update(self, request, *args, **kwargs):
        """
        Nadpisana metoda częściowej aktualizacji z dodatkową obsługą błędów
        """
        kwargs['partial'] = True
        try:
            return super().partial_update(request, *args, **kwargs)
        except Exception:
            # Usunięto logowanie błędu
            return Response(
                {'detail': 'Nie udało się zaktualizować zapotrzebowania.'},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=True, methods=['patch'])
    def change_status(self, request, pk=None):
        """
        Dedykowany endpoint do zmiany statusu
        """
        try:
            requisition = self.get_object()
            new_status = request.data.get('status')

            if not new_status:
                return Response(
                    {'detail': 'Status jest wymagany'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            valid_statuses = [status[0] for status in Requisition.REQUISITION_STATUS_CHOICES]

            if new_status not in valid_statuses:
                return Response(
                    {'detail': 'Nieprawidłowy status'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            requisition.status = new_status
            requisition.updated_by = request.user
            requisition.save()

            serializer = self.get_serializer(requisition)
            return Response(serializer.data)

        except Exception:
            # Usunięto logowanie błędu
            return Response(
                {'detail': 'Nie udało się zmienić statusu'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class RequisitionItemViewSet(viewsets.ModelViewSet):
    """API endpoint dla pozycji zapotrzebowań"""
    queryset = RequisitionItem.objects.all()
    serializer_class = RequisitionItemSerializer
    permission_classes = [permissions.IsAuthenticated, HasModulePrivilege]
    required_privilege = 'manage_requisitions'

    def update(self, request, *args, **kwargs):
        """Nadpisana metoda update z dodatkowymi logami"""
        print(f"PATCH/PUT - Update request received: {request.data}")

        # Zapisz dane przed aktualizacją
        instance = self.get_object()
        old_status = instance.status

        # Wywołaj oryginalną metodę update
        response = super().update(request, *args, **kwargs)

        # Sprawdź, czy status się zmienił i zaloguj to
        instance.refresh_from_db()
        print(f"Status change: {old_status} -> {instance.status}")
        print(f"Response data: {response.data}")

        return response

    def partial_update(self, request, *args, **kwargs):
        """Nadpisana metoda partial_update z dodatkowymi logami"""
        print(f"PATCH - Partial update request received: {request.data}")

        # Zapisz dane przed aktualizacją
        instance = self.get_object()
        old_status = instance.status

        # Wywołaj oryginalną metodę partial_update
        response = super().partial_update(request, *args, **kwargs)

        # Sprawdź, czy status się zmienił i zaloguj to
        instance.refresh_from_db()
        print(f"Status change: {old_status} -> {instance.status}")
        print(f"Response data: {response.data}")

        return response

from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
import datetime

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def export_requisitions(request):
    """Eksport zapotrzebowań do pliku Excel"""
    # Parametry filtrowania
    requisition_type = request.GET.get('type', 'material')
    status = request.GET.get('status', None)
    date_from = request.GET.get('date_from', None)
    date_to = request.GET.get('date_to', None)
    project_id = request.GET.get('project_id', None)

    # Filtrowanie zapotrzebowań
    queryset = Requisition.objects.filter(requisition_type=requisition_type).order_by('-created_at')

    # Filtruj po statusie
    if status and status != 'all':
        queryset = queryset.filter(status=status)

    # Filtruj po dacie utworzenia
    if date_from:
        try:
            date_from = datetime.datetime.strptime(date_from, '%Y-%m-%d')
            queryset = queryset.filter(created_at__gte=date_from)
        except (ValueError, TypeError):
            pass

    if date_to:
        try:
            date_to = datetime.datetime.strptime(date_to, '%Y-%m-%d')
            date_to = date_to.replace(hour=23, minute=59, second=59)
            queryset = queryset.filter(created_at__lte=date_to)
        except (ValueError, TypeError):
            pass

    # Filtruj po projekcie
    if project_id:
        queryset = queryset.filter(project_id=project_id)

    # Tworzenie pliku Excel
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Zapotrzebowania"

    # Styl nagłówków
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Nagłówki
    headers = [
        'Numer', 'Projekt', 'Status', 'Termin realizacji',
        'Data utworzenia', 'Utworzony przez', 'Wartość', 'Komentarz'
    ]

    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Szerokości kolumn
    column_widths = [15, 25, 15, 15, 15, 20, 15, 40]
    for i, width in enumerate(column_widths, 1):
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # Dodawanie danych
    status_map = {
        'to_accept': 'Do akceptacji',
        'accepted': 'Zaakceptowano',
        'rejected': 'Odrzucono',
        'in_progress': 'W trakcie realizacji',
        'completed': 'Zrealizowano'
    }

    for row_num, req in enumerate(queryset, 2):
        # Oblicz całkowitą wartość
        total_value = sum(
            item.price * item.quantity for item in req.items.all() if item.price
        )

        # Zapełnij wiersz danymi
        worksheet.cell(row=row_num, column=1).value = req.number
        worksheet.cell(row=row_num, column=2).value = req.project.name if req.project else '-'
        worksheet.cell(row=row_num, column=3).value = status_map.get(req.status, req.status)
        worksheet.cell(row=row_num, column=4).value = req.deadline.strftime('%Y-%m-%d') if req.deadline else '-'
        worksheet.cell(row=row_num, column=5).value = req.created_at.strftime('%Y-%m-%d') if req.created_at else '-'
        worksheet.cell(row=row_num, column=6).value = f"{req.created_by.first_name} {req.created_by.last_name}".strip() if req.created_by else '-'
        worksheet.cell(row=row_num, column=7).value = float(total_value)
        worksheet.cell(row=row_num, column=8).value = req.comment or '-'

        # Ustaw obramowanie dla wszystkich komórek
        for col_num in range(1, len(headers) + 1):
            worksheet.cell(row=row_num, column=col_num).border = thin_border

    # Zapisz do tymczasowego pliku
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"zapotrzebowania_{datetime.datetime.now().strftime('%Y-%m-%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)

    return response

@api_view(['POST'])
@ensure_csrf_cookie
@permission_classes([permissions.IsAuthenticated])
def validate_requisition(request):
    """Walidacja zapotrzebowania przed zapisem"""
    data = request.data
    # Sprawdź, czy projekt istnieje
    if not data.get('project'):
        return Response({
            'valid': False,
            'message': 'Wybierz projekt'
        }, status=status.HTTP_400_BAD_REQUEST)

    # Sprawdź, czy są pozycje
    if not data.get('items') or len(data.get('items', [])) == 0:
        return Response({
            'valid': False,
            'message': 'Dodaj co najmniej jedną pozycję zapotrzebowania'
        }, status=status.HTTP_400_BAD_REQUEST)

    # Sprawdź pozycje zamówienia
    for item in data.get('items', []):
        if not item.get('item'):
            return Response({
                'valid': False,
                'message': 'Wybierz przedmiot w każdej pozycji'
            }, status=status.HTTP_400_BAD_REQUEST)

        if not item.get('quantity') or int(item.get('quantity', 0)) <= 0:
            return Response({
                'valid': False,
                'message': 'Ilość przedmiotu musi być dodatnia'
            }, status=status.HTTP_400_BAD_REQUEST)

        if not item.get('price') or float(item.get('price', 0)) <= 0:
            return Response({
                'valid': False,
                'message': 'Cena przedmiotu musi być dodatnia'
            }, status=status.HTTP_400_BAD_REQUEST)

    return Response({
        'valid': True,
        'message': 'Dane zapotrzebowania są poprawne'
    })

# Add this to api/views.py

class QuarterViewSet(viewsets.ModelViewSet):
    """API endpoint for Quarters"""
    queryset = Quarter.objects.all()
    serializer_class = QuarterSerializer
    permission_classes = [permissions.IsAuthenticated, HasModulePrivilege]
    required_privilege = 'manage_quarters'  # You can define this privilege

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user, updated_by=self.request.user)

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)

# Add these additional methods to help with quarter assignments

@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def assign_employee_to_quarter(request):
    """Assign an employee to a quarter"""
    employee_id = request.data.get('employee_id')
    quarter_id = request.data.get('quarter_id')

    if not employee_id or not quarter_id:
        return Response({
            'success': False,
            'message': 'Both employee_id and quarter_id are required'
        }, status=status.HTTP_400_BAD_REQUEST)

    try:
        employee = Employee.objects.get(id=employee_id)
        quarter = Quarter.objects.get(id=quarter_id)

        # Check if quarter has space
        if quarter.employees.count() >= quarter.max_occupants:
            return Response({
                'success': False,
                'message': f'Quarter {quarter.name} is already at maximum capacity ({quarter.max_occupants} occupants)'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Assign employee to quarter
        employee.quarter = quarter
        employee.save()

        return Response({
            'success': True,
            'message': f'Employee {employee.first_name} {employee.last_name} assigned to {quarter.name}'
        })

    except Employee.DoesNotExist:
        return Response({
            'success': False,
            'message': 'Employee not found'
        }, status=status.HTTP_404_NOT_FOUND)

    except Quarter.DoesNotExist:
        return Response({
            'success': False,
            'message': 'Quarter not found'
        }, status=status.HTTP_404_NOT_FOUND)

@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def remove_employee_from_quarter(request):
    """Remove an employee from a quarter"""
    employee_id = request.data.get('employee_id')

    if not employee_id:
        return Response({
            'success': False,
            'message': 'employee_id is required'
        }, status=status.HTTP_400_BAD_REQUEST)

    try:
        employee = Employee.objects.get(id=employee_id)

        if not employee.quarter:
            return Response({
                'success': False,
                'message': f'Employee {employee.first_name} {employee.last_name} is not assigned to any quarter'
            }, status=status.HTTP_400_BAD_REQUEST)

        quarter_name = employee.quarter.name
        employee.quarter = None
        employee.save()

        return Response({
            'success': True,
            'message': f'Employee {employee.first_name} {employee.last_name} removed from {quarter_name}'
        })

    except Employee.DoesNotExist:
        return Response({
            'success': False,
            'message': 'Employee not found'
        }, status=status.HTTP_404_NOT_FOUND)

@method_decorator(ensure_csrf_cookie, name='dispatch')
class QuarterImageViewSet(viewsets.ModelViewSet):
    """API endpoint dla zdjęć kwater"""
    queryset = QuarterImage.objects.all()
    serializer_class = QuarterImageSerializer
    permission_classes = [permissions.IsAuthenticated, HasModulePrivilege]
    required_privilege = 'manage_quarters'  # To samo uprawnienie co dla kwater
    parser_classes = [parsers.MultiPartParser, parsers.FormParser]

    def get_queryset(self):
        """Filtruje zdjęcia po kwaterze, jeśli podano parametr quarter_id"""
        queryset = super().get_queryset()
        quarter_id = self.request.query_params.get('quarter_id')
        if quarter_id:
            queryset = queryset.filter(quarter_id=quarter_id)
        return queryset

    def perform_create(self, serializer):
        """Dodaje bieżącego użytkownika jako created_by"""
        serializer.save(created_by=self.request.user)

    def get_serializer_context(self):
        """Dodaje request do kontekstu serializera, aby móc generować pełne URL-e"""
        context = super().get_serializer_context()
        context['request'] = self.request
        return context

class UserSettingsViewSet(viewsets.ModelViewSet):
    """API endpoint dla ustawień użytkownika"""
    queryset = UserSettings.objects.all()
    serializer_class = UserSettingsSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        """Filtrowanie ustawień użytkownika"""
        user = self.request.user

        # Admin widzi wszystkie ustawienia
        if user.is_staff:
            return UserSettings.objects.all()

        # Zwykły użytkownik widzi tylko swoje ustawienia
        return UserSettings.objects.filter(user=user)

    def create(self, request, *args, **kwargs):
        # Sprawdź, czy ustawienia już istnieją
        existing_settings = UserSettings.objects.filter(user=request.user).first()
        if existing_settings:
            serializer = self.get_serializer(existing_settings)
            return Response(serializer.data, status=status.HTTP_200_OK)
        return super().create(request, *args, **kwargs)

    def perform_create(self, serializer):
        # Jeśli nie istnieją, utwórz nowe
        serializer.save(user=self.request.user)

    def update(self, request, *args, **kwargs):
        """Rozszerzona metoda aktualizacji, która waliduje przypisania projektu"""
        partial = kwargs.pop('partial', False)
        instance = self.get_object()

        # Upewnij się, że to są ustawienia zalogowanego użytkownika
        if instance.user != request.user and not request.user.is_staff:
            return Response(
                {'detail': 'Nie masz uprawnień do edycji ustawień innych użytkowników.'},
                status=status.HTTP_403_FORBIDDEN
            )

        # Jeśli aktualizujemy projekt
        if 'project' in request.data:
            project_id = request.data.get('project')

            # Jeśli project_id jest pusty lub null, ustawiamy na None
            if not project_id:
                request.data['project'] = None
            else:
                # Sprawdź, czy projekt istnieje
                try:
                    project = Project.objects.get(id=project_id)
                except Project.DoesNotExist:
                    return Response(
                        {'detail': f'Projekt o ID {project_id} nie istnieje.'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        # Aktualizuj pracowników w brygadzie, aby mieli ten sam projekt
        new_project = serializer.instance.project
        if new_project:
            # Pobierz wszystkich pracowników z brygady
            brigade_members = BrigadeMember.objects.filter(brigade_leader=request.user)
            for member in brigade_members:
                member.employee.current_project = new_project
                member.employee.save()

        return Response(serializer.data)

class BrigadeMemberViewSet(viewsets.ModelViewSet):
    """API endpoint dla członków brygady"""
    queryset = BrigadeMember.objects.all()
    serializer_class = BrigadeMemberSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        """Filtrowanie członków brygady"""
        user = self.request.user

        # Admin widzi wszystkich członków brygad
        if user.is_staff:
            return BrigadeMember.objects.all()

        # Zwykły użytkownik widzi tylko członków swojej brygady
        return BrigadeMember.objects.filter(brigade_leader=user)

    def perform_create(self, serializer):
        """Dodaje aktualnego użytkownika jako szefa brygady przy tworzeniu nowego członka"""
        serializer.save(brigade_leader=self.request.user)

    def create(self, request, *args, **kwargs):
        """Rozszerzenie metody create, aby sprawdzać dostępność pracownika"""
        # Sprawdź, czy pracownik jest już przypisany do innej brygady
        employee_id = request.data.get('employee')
        if BrigadeMember.objects.filter(employee_id=employee_id).exists():
            return Response(
                {'detail': 'Ten pracownik jest już przypisany do innej brygady.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        return super().create(request, *args, **kwargs)

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def my_user_settings(request):
    """Endpoint zwracający ustawienia zalogowanego użytkownika"""
    try:
        # Znajdź ustawienia użytkownika lub utwórz nowe
        user_settings, created = UserSettings.objects.get_or_create(
            user=request.user,
            defaults={'project': None}
        )

        # Dodaj informacje o utworzeniu
        response_data = {
            'created': created,
            'id': user_settings.id,
            'user': user_settings.user.id,
            'username': user_settings.user.username,
            'project': user_settings.project.id if user_settings.project else None,
            'created_at': user_settings.created_at,
            'updated_at': user_settings.updated_at,
        }

        # Dodaj informacje o projekcie, jeśli istnieje
        if user_settings.project:
            project = user_settings.project
            response_data['project_details'] = {
                'id': project.id,
                'name': project.name,
                'status': project.status,
                'status_display': project.get_status_display()
            }

        return Response(response_data)
    except Exception as e:
        return Response(
            {'detail': f"Błąd pobierania ustawień: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

# Widok do pobierania dostępnych pracowników (nie przypisanych do brygad)
@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def available_employees(request):
    """Endpoint zwracający pracowników, którzy nie są przypisani do żadnej brygady"""
    try:
        # Pobranie ID wszystkich pracowników już przypisanych do brygad
        assigned_employees = BrigadeMember.objects.values_list('employee_id', flat=True)

        # Pobranie wszystkich dostępnych pracowników
        employees = Employee.objects.exclude(id__in=assigned_employees)

        serializer = EmployeeSerializer(employees, many=True)
        return Response(serializer.data)
    except Exception as e:
        return Response({'detail': str(e)}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def update_employee_project(request):
    """Update an employee's project assignment"""
    employee_id = request.data.get('employee_id')
    project_id = request.data.get('project_id')  # Can be None to clear the project

    if not employee_id:
        return Response({
            'success': False,
            'message': 'Employee ID is required'
        }, status=status.HTTP_400_BAD_REQUEST)

    try:
        employee = Employee.objects.get(id=employee_id)

        # Update the employee's project
        if project_id:
            try:
                project = Project.objects.get(id=project_id)
                employee.current_project = project
            except Project.DoesNotExist:
                return Response({
                    'success': False,
                    'message': 'Project not found'
                }, status=status.HTTP_404_NOT_FOUND)
        else:
            # Clear the project assignment
            employee.current_project = None

        employee.save()

        return Response({
            'success': True,
            'message': f"Employee {employee.first_name} {employee.last_name}'s project updated successfully"
        })

    except Employee.DoesNotExist:
        return Response({
            'success': False,
            'message': 'Employee not found'
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({
            'success': False,
            'message': f'Error updating employee project: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def create_user_settings(request):
    """Create user settings if they don't exist"""
    try:
        # Sprawdź, czy istnieją już ustawienia dla tego użytkownika
        try:
            settings = UserSettings.objects.get(user=request.user)
            # Jeśli istnieją, zwróć je
            serializer = UserSettingsSerializer(settings)
            return Response(serializer.data)
        except UserSettings.DoesNotExist:
            # Jeśli nie istnieją, utwórz nowe bez przypisanego projektu
            settings = UserSettings.objects.create(
                user=request.user,
                project=None
            )
            serializer = UserSettingsSerializer(settings)
            return Response(serializer.data)
    except Exception as e:
        return Response(
            {'detail': f"Błąd podczas tworzenia ustawień: {str(e)}"},
            status=status.HTTP_400_BAD_REQUEST
        )

class ProgressReportViewSet(viewsets.ModelViewSet):
    """API endpoint dla raportów postępu"""
    queryset = ProgressReport.objects.all()
    serializer_class = ProgressReportSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        """Filtrowanie raportów - użytkownik widzi tylko swoje raporty lub wszystkie, jeśli jest adminem"""
        user = self.request.user

        # Admin widzi wszystkie raporty
        if user.is_staff:
            return ProgressReport.objects.all().order_by('-date')

        # Pozostali użytkownicy widzą tylko swoje raporty
        return ProgressReport.objects.filter(created_by=user).order_by('-date')

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

class ProgressReportEntryViewSet(viewsets.ModelViewSet):
    """API endpoint dla wpisów w raportach postępu"""
    queryset = ProgressReportEntry.objects.all()
    serializer_class = ProgressReportEntrySerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        """Filtrowanie wpisów - powiązanie z raportem"""
        report_id = self.request.query_params.get('report_id')
        if report_id:
            return ProgressReportEntry.objects.filter(report_id=report_id)
        return ProgressReportEntry.objects.all()

class ProgressReportImageViewSet(viewsets.ModelViewSet):
    """API endpoint dla zdjęć raportów postępu"""
    queryset = ProgressReportImage.objects.all()
    serializer_class = ProgressReportImageSerializer
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [parsers.MultiPartParser, parsers.FormParser]

    def get_queryset(self):
        """Filtrowanie zdjęć po raporcie, jeśli podano parametr report_id"""
        queryset = super().get_queryset()
        report_id = self.request.query_params.get('report_id')
        if report_id:
            queryset = queryset.filter(report_id=report_id)
        return queryset

    def perform_create(self, serializer):
        """Dodaje bieżącego użytkownika jako created_by"""
        serializer.save(created_by=self.request.user)

    def get_serializer_context(self):
        """Dodaje request do kontekstu serializera, aby móc generować pełne URL-e"""
        context = super().get_serializer_context()
        context['request'] = self.request
        return context

@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def create_progress_report(request):
    """Endpoint do tworzenia raportu postępu wraz z wpisami w jednym żądaniu"""
    try:
        # Dane raportu
        report_data = {
            'date': request.data.get('date'),
            'project': request.data.get('project'),
            'is_draft': request.data.get('is_draft', False)  # Dodane obsługa flagi draft
        }

        # Walidacja
        if not report_data['date'] or not report_data['project']:
            return Response(
                {'detail': 'Data i projekt są wymagane'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Sprawdź, czy już istnieje raport z tą datą i projektem dla tego użytkownika
        existing_report = ProgressReport.objects.filter(
            date=report_data['date'],
            project=report_data['project'],
            created_by=request.user
        ).first()

        if existing_report:
            # Zaktualizuj status draft istniejącego raportu
            existing_report.is_draft = report_data['is_draft']
            existing_report.save()

            # Zwróć istniejący raport
            serializer = ProgressReportSerializer(existing_report)
            return Response(serializer.data)

        # Utwórz raport
        report = ProgressReport.objects.create(
            date=report_data['date'],
            project_id=report_data['project'],
            created_by=request.user,
            is_draft=report_data['is_draft']  # Zapisz status draft
        )

        # Utwórz wpisy
        entries_data = request.data.get('entries', [])
        for entry_data in entries_data:
            if entry_data.get('employee') and entry_data.get('hours_worked') is not None:
                ProgressReportEntry.objects.create(
                    report=report,
                    employee_id=entry_data['employee'],
                    hours_worked=entry_data['hours_worked'],
                    notes=entry_data.get('notes', '')
                )

        # Zwróć utworzony raport z wpisami
        serializer = ProgressReportSerializer(report)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    except Exception as e:
        return Response(
            {'detail': str(e)},
            status=status.HTTP_400_BAD_REQUEST
        )

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def get_progress_reports_for_date(request):
    """Zwraca raporty postępu dla określonej daty i użytkownika"""
    try:
        date = request.query_params.get('date')
        if not date:
            return Response(
                {'detail': 'Data jest wymagana'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Pobierz raporty dla danej daty
        reports = ProgressReport.objects.filter(
            date=date,
            created_by=request.user
        )

        serializer = ProgressReportSerializer(reports, many=True)
        return Response(serializer.data)

    except Exception as e:
        return Response(
            {'detail': str(e)},
            status=status.HTTP_400_BAD_REQUEST
        )

class HRRequisitionPositionViewSet(viewsets.ModelViewSet):
    """API endpoint dla pozycji zapotrzebowań HR"""
    queryset = HRRequisitionPosition.objects.all()
    serializer_class = HRRequisitionPositionSerializer
    permission_classes = [permissions.IsAuthenticated, HasModulePrivilege]
    required_privilege = 'manage_hr_requisitions'

class HRRequisitionViewSet(viewsets.ModelViewSet):
    """API endpoint dla zapotrzebowań HR"""
    queryset = HRRequisition.objects.all()
    serializer_class = HRRequisitionSerializer
    permission_classes = [permissions.IsAuthenticated, HasModulePrivilege]
    required_privilege = 'manage_hr_requisitions'

    def get_queryset(self):
        """Filtruj zapotrzebowania z obsługą wyszukiwania"""
        user = self.request.user
        # Użytkownicy z uprawnieniem 'view_all_requisitions' lub admin mogą widzieć wszystkie zapotrzebowania
        if user.is_staff or hasattr(user, 'profile') and user.profile.has_privilege('view_all_requisitions'):
            queryset = HRRequisition.objects.all().order_by('-created_at')
        else:
            # Pozostali użytkownicy widzą tylko swoje zapotrzebowania
            queryset = HRRequisition.objects.filter(created_by=user).order_by('-created_at')

        # Filtruj po frazie wyszukiwania
        search_term = self.request.query_params.get('search', None)
        if search_term:
            # Wyszukiwanie w podstawowych polach zapotrzebowania
            queryset = queryset.filter(
                Q(number__icontains=search_term) |
                Q(comment__icontains=search_term) |
                Q(special_requirements__icontains=search_term) |
                Q(project__name__icontains=search_term)
            )

        return queryset.distinct()

    def get_serializer_context(self):
        """Dodaj request do kontekstu serializera, aby mieć dostęp do aktualnego użytkownika"""
        context = super().get_serializer_context()
        context['request'] = self.request
        return context

    def perform_create(self, serializer):
        """
        Proces tworzenia zapotrzebowania HR
        """
        requisition = serializer.save(
            created_by=self.request.user,
            updated_by=self.request.user
        )

    def perform_update(self, serializer):
        """
        Obsługa aktualizacji zapotrzebowania HR
        """
        # Pobierz aktualny stan przed aktualizacją
        instance = self.get_object()
        old_status = instance.status

        # Zaktualizuj zapotrzebowanie
        updated_instance = serializer.save(
            updated_by=self.request.user
        )

    @action(detail=True, methods=['patch'])
    def change_status(self, request, pk=None):
        """
        Dedykowany endpoint do zmiany statusu
        """
        try:
            requisition = self.get_object()
            new_status = request.data.get('status')

            if not new_status:
                return Response(
                    {'detail': 'Status jest wymagany'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            valid_statuses = [status_choice[0] for status_choice in Requisition.REQUISITION_STATUS_CHOICES]

            if new_status not in valid_statuses:
                return Response(
                    {'detail': 'Nieprawidłowy status'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            requisition.status = new_status
            requisition.updated_by = request.user
            requisition.save()

            serializer = self.get_serializer(requisition)
            return Response(serializer.data)

        except Exception as e:
            return Response(
                {'detail': f'Nie udało się zmienić statusu: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def validate_hr_requisition(request):
    """Walidacja zapotrzebowania HR przed zapisem"""
    data = request.data

    # Sprawdź, czy projekt istnieje
    if not data.get('project'):
        return Response({
            'valid': False,
            'message': 'Wybierz projekt'
        }, status=status.HTTP_400_BAD_REQUEST)

    # Sprawdź, czy podano termin realizacji
    if not data.get('deadline'):
        return Response({
            'valid': False,
            'message': 'Określ termin realizacji'
        }, status=status.HTTP_400_BAD_REQUEST)

    # Sprawdź, czy podano doświadczenie
    if not data.get('experience'):
        return Response({
            'valid': False,
            'message': 'Określ wymagane doświadczenie'
        }, status=status.HTTP_400_BAD_REQUEST)

    # Sprawdź, czy są stanowiska
    if not data.get('positions') or len(data.get('positions', [])) == 0:
        return Response({
            'valid': False,
            'message': 'Dodaj co najmniej jedno stanowisko'
        }, status=status.HTTP_400_BAD_REQUEST)

    # Sprawdź pozycje stanowiska
    for position in data.get('positions', []):
        if not position.get('position'):
            return Response({
                'valid': False,
                'message': 'Wybierz stanowisko w każdej pozycji'
            }, status=status.HTTP_400_BAD_REQUEST)

        if not position.get('quantity') or int(position.get('quantity', 0)) <= 0:
            return Response({
                'valid': False,
                'message': 'Ilość pracowników musi być dodatnia'
            }, status=status.HTTP_400_BAD_REQUEST)

    return Response({
        'valid': True,
        'message': 'Dane zapotrzebowania HR są poprawne'
    })

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def get_employee_by_tag(request, tag_id):
    """Endpoint zwracający pracownika po ID tagu NFC"""
    try:
        # Znajdź tag pracownika o podanym numerze seryjnym
        employee_tag = Empl_tag.objects.filter(serial=tag_id).first()

        if not employee_tag:
            return Response(
                {'detail': 'Nie znaleziono tagu o podanym ID'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Znajdź pracownika powiązanego z tym tagiem
        employee = Employee.objects.filter(employee_tag=employee_tag).first()

        if not employee:
            return Response(
                {'detail': 'Nie znaleziono pracownika z tym tagiem'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Serializuj i zwróć dane pracownika
        serializer = EmployeeSerializer(employee)
        return Response(serializer.data)

    except Exception as e:
        return Response(
            {'detail': f'Wystąpił błąd: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

@method_decorator(ensure_csrf_cookie, name='dispatch')
class TransportRequestViewSet(viewsets.ModelViewSet):
    """API endpoint dla zapotrzebowań na transport"""
    queryset = TransportRequest.objects.all()
    serializer_class = TransportRequestSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        """Filtruje zapotrzebowania na transport w zależności od uprawnień użytkownika"""
        user = self.request.user

        # Adminom pokazujemy wszystkie zapotrzebowania
        if user.is_staff or hasattr(user, 'profile') and user.profile.has_privilege('view_all_transports'):
            return TransportRequest.objects.all().order_by('-created_at')

        # Pozostali użytkownicy widzą tylko swoje zapotrzebowania
        return TransportRequest.objects.filter(created_by=user).order_by('-created_at')

    def create(self, request, *args, **kwargs):
        """Utworzenie nowego zapotrzebowania transportowego wraz z przesyłkami"""
        # Ekstrakcja danych przesyłek z żądania
        items_data = request.data.pop('items', [])

        # Serializacja danych zapotrzebowania
        serializer = self.get_serializer(
            data=request.data,
            context={'request': request, 'items': items_data}
        )

        # Walidacja danych
        serializer.is_valid(raise_exception=True)

        # Tworzenie zapotrzebowania i przesyłek
        self.perform_create(serializer)

        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def perform_create(self, serializer):
        serializer.save(
            created_by=self.request.user,
            updated_by=self.request.user
        )

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)

    @action(detail=True, methods=['post'])
    def change_status(self, request, pk=None):
        """Dedykowany endpoint do zmiany statusu zapotrzebowania"""
        transport = self.get_object()
        new_status = request.data.get('status')

        if not new_status:
            return Response(
                {'detail': 'Status jest wymagany'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Sprawdź czy nowy status jest dopuszczalny
        valid_statuses = [status[0] for status in TransportRequest.STATUS_CHOICES]

        if new_status not in valid_statuses:
            return Response(
                {'detail': 'Nieprawidłowy status'},
                status=status.HTTP_400_BAD_REQUEST
            )

        transport.status = new_status
        transport.updated_by = request.user
        transport.save()

        serializer = self.get_serializer(transport)
        return Response(serializer.data)

class TransportItemViewSet(viewsets.ModelViewSet):
    """API endpoint dla przesyłek w transporcie"""
    queryset = TransportItem.objects.all()
    serializer_class = TransportItemSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        """Filtruje przesyłki po powiązanym transporcie"""
        transport_id = self.request.query_params.get('transport_id')
        if transport_id:
            return TransportItem.objects.filter(transport_id=transport_id)
        return TransportItem.objects.all()

@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def validate_transport(request):
    """Walidacja formularza transportu przed zapisaniem"""
    data = request.data

    # Sprawdź podstawowe dane
    errors = {}

    # Sprawdź dane załadunku
    if not data.get('loading_project') and not data.get('loading_other_address'):
        errors['loading'] = 'Musisz podać projekt lub inny adres załadunku'

    if not data.get('loading_date'):
        errors['loading_date'] = 'Data załadunku jest wymagana'

    # Sprawdź dane rozładunku
    if not data.get('unloading_project') and not data.get('unloading_other_address'):
        errors['unloading'] = 'Musisz podać projekt lub inny adres rozładunku'

    if not data.get('unloading_date'):
        errors['unloading_date'] = 'Data rozładunku jest wymagana'

    # Sprawdź projekt kosztowy
    if not data.get('cost_project'):
        errors['cost_project'] = 'Projekt kosztowy jest wymagany'

    # Sprawdź czy są przedmioty w transporcie
    if not data.get('items') or len(data.get('items', [])) == 0:
        errors['items'] = 'Dodaj co najmniej jeden przedmiot do transportu'

    if errors:
        return Response({'valid': False, 'errors': errors}, status=status.HTTP_400_BAD_REQUEST)

    return Response({'valid': True}, status=status.HTTP_200_OK)

class ProjectActivityConfigViewSet(viewsets.ModelViewSet):
    """API endpoint dla konfiguracji aktywności projektu"""
    queryset = ProjectActivityConfig.objects.all()
    serializer_class = ProjectActivityConfigSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        """Filtrowanie konfiguracji aktywności"""
        project_id = self.request.query_params.get('project_id', None)
        if project_id:
            return ProjectActivityConfig.objects.filter(project_id=project_id)
        return ProjectActivityConfig.objects.all()

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user, updated_by=self.request.user)

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)

class ProgressReportActivityViewSet(viewsets.ModelViewSet):
    """API endpoint dla aktywności w raportach postępu"""
    queryset = ProgressReportActivity.objects.all()
    serializer_class = ProgressReportActivitySerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        """Filtrowanie aktywności po raporcie"""
        report_id = self.request.query_params.get('report_id', None)
        if report_id:
            return ProgressReportActivity.objects.filter(report_id=report_id)
        return ProgressReportActivity.objects.all()

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def get_project_activities_config(request):
    """Pobranie konfiguracji aktywności dla projektu"""
    project_id = request.query_params.get('project_id', None)

    if not project_id:
        return Response(
            {'detail': 'Identyfikator projektu jest wymagany'},
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        # Pobierz konfigurację z bazy danych
        config = ProjectActivityConfig.objects.filter(project_id=project_id).first()

        if config:
            serializer = ProjectActivityConfigSerializer(config)
            return Response(serializer.data)

        # Jeśli nie znaleziono konfiguracji, zwróć pustą odpowiedź
        return Response({
            'project': project_id,
            'config_data': None
        })

    except Exception as e:
        return Response(
            {'detail': str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def upload_project_activities_config(request):
    """Wgranie konfiguracji aktywności dla projektu"""
    project_id = request.data.get('project_id')
    config_file = request.FILES.get('config_file')

    if not project_id or not config_file:
        return Response(
            {'detail': 'Identyfikator projektu i plik konfiguracyjny są wymagane'},
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        # Sprawdź czy projekt istnieje
        try:
            project = Project.objects.get(id=project_id)
        except Project.DoesNotExist:
            return Response(
                {'detail': 'Projekt nie istnieje'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Wczytaj plik JSON
        try:
            json_data = json.load(config_file)
        except json.JSONDecodeError:
            return Response(
                {'detail': 'Nieprawidłowy format pliku JSON'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Utwórz lub zaktualizuj konfigurację
        config, created = ProjectActivityConfig.objects.update_or_create(
            project=project,
            defaults={
                'config_data': json_data,
                'created_by': request.user if created else None,
                'updated_by': request.user
            }
        )

        serializer = ProjectActivityConfigSerializer(config)
        return Response(serializer.data)

    except Exception as e:
        return Response(
            {'detail': str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def add_activities_to_report(request):
    """Endpoint do dodawania aktywności do raportu postępu"""
    try:
        report_id = request.data.get('report_id')
        activities_data = request.data.get('activities', [])

        if not report_id:
            return Response({'detail': 'Brak ID raportu'}, status=status.HTTP_400_BAD_REQUEST)

        # Pobierz raport
        try:
            report = ProgressReport.objects.get(id=report_id)
        except ProgressReport.DoesNotExist:
            return Response({'detail': 'Raport nie istnieje'}, status=status.HTTP_404_NOT_FOUND)

        # Usuń istniejące aktywności dla tego raportu
        ProgressReportActivity.objects.filter(report=report).delete()

        # Dodaj nowe aktywności
        for activity_data in activities_data:
            activity_data['report'] = report_id
            serializer = ProgressReportActivitySerializer(data=activity_data)
            if serializer.is_valid():
                serializer.save()
            else:
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        return Response({'detail': 'Aktywności zapisane pomyślnie'}, status=status.HTTP_201_CREATED)

    except Exception as e:
        return Response({'detail': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)